from pathlib import Path
import re
import pymongo
import sys
import base64

# Optional Excel parser
try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

# ---------- CONFIG ----------
BASE_DIR = Path(r"D:/migratetomongo/AutomationTool/src/test/resources/testData")
MONGO_URI = "mongodb://localhost:27017/"
DB_NAME = "te"
DO_INSERT = True   # Set False to dry-run (no DB writes)
# ----------------------------

def sanitize_collection_name(name: str) -> str:
    """Replace invalid characters with underscore so the collection name is Mongo-safe."""
    return re.sub(r'[^\w\-]', "_", name)

def ensure_unique_index(collection):
    """Create a unique index on (category, orderLine, apiName) to avoid duplicates."""
    try:
        collection.create_index(
            [("category", pymongo.ASCENDING),
             ("orderLine", pymongo.ASCENDING),
             ("apiName", pymongo.ASCENDING)],
            unique=True,
            name="uq_category_orderline_apiname"
        )
    except Exception as e:
        print(f"  ⚠️ Could not create unique index on {collection.name}: {e}")

def parse_xlsx_to_json(filepath: Path):
    """
    Parse an .xlsx file into a dict: { sheet_name: [ {col: value, ...}, ... ], ... }
    Uses first row as header. If openpyxl not available, raises ImportError.
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl not available")
    wb = load_workbook(filename=str(filepath), data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            result[sheet_name] = []
            continue
        # first row as header
        header = list(rows[0])
        # sanitize header names (empty -> col_<idx>)
        header_clean = []
        for idx, h in enumerate(header):
            if h is None or str(h).strip() == "":
                header_clean.append(f"col_{idx}")
            else:
                header_clean.append(str(h).strip())
        data_rows = []
        for r in rows[1:]:
            row_dict = {}
            for col_idx, cell in enumerate(r):
                col_name = header_clean[col_idx] if col_idx < len(header_clean) else f"col_{col_idx}"
                row_dict[col_name] = cell
            data_rows.append(row_dict)
        result[sheet_name] = data_rows
    return result

def file_to_base64(filepath: Path):
    b = filepath.read_bytes()
    return base64.b64encode(b).decode("ascii")

def insert_doc(collection, doc, file_name, collection_name, orderLine, dry_run=False):
    """
    Insert doc only if it doesn't already exist in collection.
    - collection: pymongo Collection object or None (if DB not available)
    - dry_run: True => simulate (no writes)
    Returns: True (inserted or would insert), False (skipped duplicate), None (error)
    """
    query = {
        "category": doc["category"],
        "orderLine": doc["orderLine"],
        "apiName": doc["apiName"]
    }

    # If collection not available, just simulate or show message
    if collection is None:
        if dry_run:
            print(f"  (DRY) Would insert {file_name} into '{collection_name}' (orderLine='{orderLine}')")
            return True
        else:
            print(f"  ⚠️ MongoDB not available. Skipping actual insert for {file_name}.")
            return None

    try:
        existing = collection.find_one(query)
    except Exception as e:
        print(f"  ❌ DB error while checking existing for {file_name}: {e}")
        return None

    if existing:
        print(f"  ⚠️ Skipped duplicate: {file_name} already exists in '{collection_name}' (orderLine='{orderLine}')")
        return False  # skipped

    if dry_run:
        print(f"  (DRY) Would insert {file_name} into '{collection_name}' (orderLine='{orderLine}')")
        return True

    try:
        collection.insert_one(doc)
        print(f"  ✅ Inserted {file_name} into '{collection_name}' (orderLine='{orderLine}')")
        return True
    except pymongo.errors.DuplicateKeyError:
        # race condition possible; treat as skipped
        print(f"  ⚠️ DuplicateKeyError (skipped) for {file_name} in '{collection_name}'")
        return False
    except Exception as e:
        print(f"  ❌ DB error inserting {file_name}: {e}")
        return None

def process_file(item: Path, category: str, orderLine: str, collection, collection_name, counters):
    """
    Read/parse a single file and attempt insertion.
    counters: dict with keys total, inserted, skipped, errors
    """
    counters["total"] += 1
    apiName = item.stem
    suffix = item.suffix.lower()
    file_name = item.name

    try:
        if suffix == ".xml":
            xml_text = item.read_text(encoding="utf-8", errors="replace")
            doc = {
                "category": category,
                "orderLine": orderLine,
                "Content": xml_text,
                "ParsedContent": None,
                "apiName": apiName,
                "fileType": "xml"
            }
        elif suffix == ".xlsx":
            if OPENPYXL_AVAILABLE:
                parsed = parse_xlsx_to_json(item)
                doc = {
                    "category": category,
                    "orderLine": orderLine,
                    "Content": None,
                    "ParsedContent": parsed,
                    "apiName": apiName,
                    "fileType": "xlsx"
                }
            else:
                b64 = file_to_base64(item)
                doc = {
                    "category": category,
                    "orderLine": orderLine,
                    "Content": None,
                    "Content_base64": b64,
                    "ParsedContent": None,
                    "apiName": apiName,
                    "fileType": "xlsx"
                }
        else:
            print(f"  Skipping unsupported file type: {file_name}")
            counters["skipped"] += 1
            return
    except Exception as e:
        print(f"  ❌ Error processing file {file_name}: {e}")
        counters["errors"] += 1
        return

    res = insert_doc(collection, doc, file_name, collection_name, orderLine, dry_run=(not DO_INSERT))
    if res is True:
        counters["inserted"] += 1
    elif res is False:
        counters["skipped"] += 1
    else:
        counters["errors"] += 1

def main():
    print(f"Starting script\nBASE_DIR = {BASE_DIR}\nMONGO_URI = {MONGO_URI}\nDB = {DB_NAME}\nDO_INSERT = {DO_INSERT}\n")
    if not OPENPYXL_AVAILABLE:
        print("Note: openpyxl not available — XLSX will be stored as base64 in 'Content_base64'.")
        print("To get parsed XLSX, run: pip install openpyxl\n")

    # Validate base dir
    if not BASE_DIR.exists() or not BASE_DIR.is_dir():
        print(f"ERROR: BASE_DIR does not exist or is not a directory: {BASE_DIR}")
        sys.exit(1)

    # Connect to MongoDB
    db = None
    try:
        client = pymongo.MongoClient(MONGO_URI, serverSelectionTimeoutMS=5000)
        client.admin.command("ping")
        db = client[DB_NAME]
        print("✅ Connected to MongoDB (ping successful).\n")
    except Exception as e:
        print(f"❌ Could not connect to MongoDB: {e}\nContinuing in DRY mode (no DB writes).")
        db = None

    counters = {"total": 0, "inserted": 0, "skipped": 0, "errors": 0}

    # Iterate immediate children of BASE_DIR -> treat each as a 'category'
    for category_path in sorted(BASE_DIR.iterdir()):
        if not category_path.is_dir():
            print(f"Skipping non-directory at top-level: {category_path.name}")
            continue

        category = category_path.name
        collection_name = sanitize_collection_name(category)
        collection = db[collection_name] if db is not None else None

        print(f"\nProcessing category folder: '{category}' -> collection '{collection_name}'")

        # Ensure unique index if DB available
        if collection is not None:
            ensure_unique_index(collection)

        # 1) Process files directly inside category folder (orderLine = "")
        for item in sorted(category_path.iterdir()):
            if item.is_file():
                process_file(item, category, "", collection, collection_name, counters)

        # 2) Process immediate subfolders (each subfolder => orderLine)
        subfolders = [p for p in sorted(category_path.iterdir()) if p.is_dir()]
        if not subfolders:
            print(f"  (Info) No subfolders under '{category}'.")
        for sub in subfolders:
            orderLine = sub.name
            print(f"  Subfolder (orderLine): '{orderLine}'")
            files_in_sub = sorted(sub.iterdir())
            if not files_in_sub:
                print(f"    ⚠️ No files found in subfolder: {sub}")
            for f in files_in_sub:
                if f.is_dir():
                    print(f"    Skipping nested directory: {f.name}")
                    continue
                process_file(f, category, orderLine, collection, collection_name, counters)

    print("\n--- Summary ---")
    print(f"Total files encountered: {counters['total']}")
    print(f"Inserted (or would insert in dry-run): {counters['inserted']}")
    print(f"Skipped (non-supported/duplicates): {counters['skipped']}")
    print(f"Errors: {counters['errors']}")
    print("Finished.")

if __name__ == "__main__":
    main()